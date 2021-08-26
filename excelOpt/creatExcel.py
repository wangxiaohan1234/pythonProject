from openpyxl import Workbook, load_workbook
wb = load_workbook("/Users/snowang/Downloads/pythontest/123.xlsx")
ws = wb["Sheet1"]
total_list = []
for row in ws.rows:
    row_list = []
    for cell in row:
        row_list.append(cell.value)
    total_list.append(row_list)
namedict = {}
valuedict = {}
# for term in total_list:
#     if term[0] is None or term[0] == "员工姓名":
#         continue
#     else:
#         namedict[term[0]] = namedict.get(term[0], 0) + 1
#         valuedict.setdefault(term[0], []).append(term[-1])
for index in range(len(total_list)):
    if index == 0:
        continue
    else:
        print(index)

newwb = Workbook()
newsheet = newwb.active
newsheet.title = "Sheet1"
# 表头
newsheet["A1"] = "123333"
newsheet.merge_cells(range_string=None, start_row=1, start_column=1, end_row=1, end_column=3)
newsheet["A2"] = "员工姓名"
newsheet["B2"] = "推荐数量"
newsheet["C2"] = "累计认购金额"

# 单元格
i = 3
for name in namedict:
    newsheet.cell(row=i, column=1).value = name
    newsheet.cell(row=i, column=2).value = namedict[name]
    newsheet.cell(row=i, column=3).value = sum(valuedict[name])
    i = i + 1
# newwb.save("/Users/snowang/Downloads/pythontest/1234.xlsx")
